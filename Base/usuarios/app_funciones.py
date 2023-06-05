from openpyxl import load_workbook
import configparser, datetime, importlib, os

from django.core.exceptions import ValidationError
from django.contrib.staticfiles.storage import staticfiles_storage
from django.db.models import Q
from django.utils.translation import gettext as _

# Instanciación dinamica de clases

class ObjetoDinamico():
	def __init__(self, pPaquete, pClase):
		'''
			Crea un objeto de la clase, con base al string del paquete en el cual se coloca
			y la clase que se desea instanciar

			PARAMETROS
			pPaquete (string): 	paquete al cual pertenece la clase que se desea instanciar
			pClase (string): 	clase que se desea instanciar
			**kwargs (dict): 	diccionario con los campos que se desea filtrar, si se quiere una instancia específica

			RETURN
			Devuelve el objeto creado
		'''
		pPaquete = importlib.import_module(pPaquete)
		self.descripcion = f'{pPaquete}/{pClase}'
		self.clase = getattr(pPaquete, pClase)
		self.obj = getattr(pPaquete, pClase)()

	def get_class(self):
		return self.clase

	def get_object(self):
		return self.obj

	def get_or_create(self, pId, campo="id", **kwargs):
		'''
			Permite filtrar de por medio del queryset, utiliando un diccionario en busqueda

			PARAMETROS
			id 					identificador del objeto a validar
			**kwargs (dict): 	diccionario con los parametros utilizandos en el filtro

			RETURN
			bool, obj 			devuelve si es nuevo el objeto o no y el objeto
		'''
		try:
			filtro = dict(zip(campo, pId))
			return False, self.clase.objects.get(**filtro)
		except Exception as e:
			self.asigna_parametros(pAutoSave=True, **kwargs)
			return True, self.obj

	def asigna_parametros(self, pAutoSave=False, **kwargs):
		'''
			Busca o crea objetos dentro del paquete qliksense.models

			PARAMETROS
			autosave (bool): 	determina si se guarda el objecto en base de datos automáticamente
			**kwargs (dict): 	diccionario con los parametros y valores a asignar

			RETURN
			devuelve el objeto con los parametros, posterior a guardarlo en base de datos
		'''
		try:
			parametros = kwargs
			for k, v in parametros.items():
				setattr(self.obj, k, v)

			if pAutoSave:
				self.obj.save()
		except Exception as e:
			print('Error in method: usuarios.app_funciones.asigna_parametros()')
			raise e

	def elimina_repetidos(self, *args, **kwargs):
		'''
			Elimina valores repetidos tomando en consideracion unicamente los campos provistos

			PARAMETROS
			*args 		lista de campos a considerar para la eliminacion
			**kwargs	valores para determinar un subgrupo de valores, (filtros)
		'''
		campos = args
		filtros = kwargs
		valores_distintos = self.clase.objects.filter(**filtros).values_list(*campos).distinct()
		for valores in valores_distintos:
			diccionario = dict(zip(campos, valores))
			ids = self.clase.objects.filter(**filtros).filter(**diccionario)#.values_list('id', flat=True)
			if len(ids)>1:
				ids.exclude(id=ids[0].id).delete()

	def valida_modificaciones(pObjeto, pAutoSave=False, **kwargs):
		'''
		Valida si un objeto Django ha tenido modificaciones o no

		PARAMETROS
		pObjeto		Objeto que se desea comparar
		kwargs		dict con los valores a comparar

		RETURN
		bool, obj 	Boolean que indica si se modificó o no y el objeto final
		'''
		parametros = kwargs
		modificado = False

		for k, v in parametros.items():
			if getattr(pObjeto, k) != v:
				setattr(pObjeto, k, v)
				modificado = True

		if modificado and pAutoSave:
			pObjeto.save()
			modificado = False

		return modificado, pObjeto


class Configuraciones():
	config = configparser.ConfigParser()

	def __init__(self, path=None, static=True, file='configuraciones.cfg'):
		'''
			Lee los archivos de configuración

			PARAMETROS
			path (string): 	ruta base de la aplicacion donde se encuentra el archivo
			static (bool): 	determina si se utilia la ruta defaul static de django
			file (string): 	nombre del archivo de configuración (con su extension)
		'''
		if path is None:
			url = './' #ruta relativa; os.getcwd() #ruta completa;

		if static: 
			url += fr'{staticfiles_storage.url(file)}'
		else:
			url += fr'{file}'

		self.config.readfp(open(fr'{url}'))

	def get_value(self, pSection, pVariable):
		'''
			Devuelve un valor específico del archivo de configuración

			PARAMETROS
			pSection (string): 	sección del archivo de donde se quiere obtener el valor
			pVariable (string): variable de la que se quiere obtener el valor
			
			RETURN
			Valor de la variable indicada
		'''
		return self.config[f'{pSection}'][f'{pVariable}']


class Archivo():
	'''
		Carga los archivos y genera un dict con la información para consumirlos desde
		otras clases validando el tipo de archvio y la parametrización de 

		  - Usuarios.models ParametriaArchivoEncabezado, ParametriaArchivoDetalle -  

		CAMPOS
		file 					(file) 		Carga el archivo y lo instancia de acuerdo con la extension
		campos_extra_en_archivo	(string) 	Campos en el archivo que son innecearios
		error_en_data 			(array) 	Errores al momento de cargar la información
		parametrizacion 		(ParametriaArchivoDetalle) Parametrización de campos (detalle archivo)
		sheet # eliminar (solo se usa en exceles)
	'''
	def __init__(self, file, parametria_extensiones, parametria_detalle):
		'''
			PARAMETROS
			file 					Archivo cargado desde el formulario
			parametria_extensiones 	Arreglo con el listado de extensiones validas para el archivo
			parametria_detalle		Objeto ParametriaArchivoDetalle (con el detalle parametrizado)

			RETURN
			None
		'''
		self.file 						= file
		self.campos_extra_en_archivo 	= None
		self.error_en_data 				= []
		self.parametrizacion 			= parametria_detalle

		if not self.file.content_type in parametria_extensiones:
			raise ValidationError(_('Archivo no valido'))

	def leer_archivo(self, hoja=None):
		if self.file.content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
			### para archivos excel
			self.leer_archivo_xlsx(hoja)
	
	def obtiene_hoja(self, hoja=None):
		if hoja:
			return load_workbook(self.file)[hoja]
		else:
			return load_workbook(self.file).active		
	
	def lee_encabezados_xlsx(self, hoja=None, linea_encabezado=1):
		errores, campos_extra, campos_final = [], [], {}
		parametrizacion = self.parametrizacion

		encabezados = [item for item in self.obtiene_hoja(hoja)[linea_encabezado]]

		for item in encabezados:
			parametro = self.parametrizacion.filter(display=item.value)
			if parametro and item.value in campos_final:
				errores.append(ValidationError(_('Campo parametrizado %(campo)s repetido en el archivo (columna: %(columna)s)'), params={'campo': item.value, 'columna': item.col_idx}))
			elif parametro:
				campos_final[item.value] = {
					'posicion': item.col_idx-1,
					'tipo': parametro[0].tipo,
					'validacion': parametro[0].validacion if parametro[0].validacion else None,
				}
				parametrizacion = parametrizacion.exclude(display=item.value)
			else: 
				campos_extra.append(f'{item.value} ({item.col_idx})')

		self.campos_extra_en_archivo = f'Campos extra en el archivo: {", ".join(campos_extra)}' if campos_extra else None

		if parametrizacion.count()>0:
			errores.append(
				ValidationError(_('Campos no parametrizados: %(campo)s'), 
				params={'campo': ', '.join(parametrizacion.values_list('display', flat=True))})
			)

		if errores:
			raise ValidationError(errores)

		return campos_final

	def leer_archivo_xlsx(self, hoja=None, linea_encabezado=1):
		linea_data, data = linea_encabezado+1, {}

		encabezados = self.lee_encabezados_xlsx(hoja, linea_encabezado)

		for index, linea in enumerate(self.obtiene_hoja(hoja).iter_rows(min_row=linea_data)):
			detalle = {}
			registro_valido = True
			for k, v in encabezados.items():
				valido, detalle[k.upper()] = self.valida_tipo_dato(linea[v['posicion']].value, v['tipo'], validacion=v['validacion'])
				if not valido:
					registro_valido = False
			if not registro_valido:
				self.error_en_data.append(f'Error en la linea: {index+linea_data} - Detalle: {detalle}')
				continue
			data[index+linea_data] = detalle
		return data

	def valida_tipo_dato(self, valor, tipo, **kwargs):
		try:
			if tipo == 'DATE':
				formato = kwargs['validacion'] if kwargs['validacion'] else '%d/%m/%Y'
				return 1, datetime.datetime.strptime(valor, formato).strftime(formato)
			elif tipo == 'INT':
				return 1, int(valor)
			elif tipo == 'DEC':
				return 1, float(valor)
			elif tipo == 'BOOL':
				return 1, valor != 0
			elif tipo == 'TEXT':
				validacion = kwargs['validacion'].upper().split(';')
				if 'BLANK=FALSE' in validacion and len(valor)==0:
					return 0, '<>'
				return 1, str(valor)
			else:
				return 0, f'<{valor}>'
		except Exception as e:
			print(e)
			return 0, f'<{valor}>'

###
### Búsqueda genérica
###
def busqueda_nombres(campos, valores):
    '''
        Permite generar una busqueda compleja de multiples valores en multiples campos
        por medio de sentencias OR
        campos: campo y tipo de consulta (ej. [nombre__icontains, apellido__icontains])
        valores: valores a buscar en la consulta OR separado por espacio (ej. pablo godoy) 
    '''
    q = Q()
    for campo in campos:
        for valor in valores:
            q |= Q(**{f'{campo}' : valor})
    return q

###
### Validación por extesniones de archivo
###

		