import importlib
from openpyxl import load_workbook

from django.conf import settings
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import (LoginView, LogoutView, PasswordResetConfirmView, 
    PasswordResetView, PasswordResetDoneView, PasswordResetCompleteView, PasswordChangeView)
from django.contrib.messages.views import SuccessMessageMixin
from django.db.models import Q
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.utils.translation import gettext as _
from django.urls import reverse_lazy

from .app_funciones import Configuraciones
from .models import Usuario, Regionalizacion
from .forms import CustomUserCreationForm, CustomUserUpdateForm, PerfilForm, RegionalizacionUploadForm
from .personal_views import (PersonalTemplateView, PersonalListView, PersonalFormView, 
    PersonalUpdateView)

#
# LECTURA DE ARCHIVO DE CONFIGURACIÓN
#

gConfiguracion = Configuraciones()

#
# FUNCIONES GENERICAS
#

def BusquedaNombres(campos, valores):
    '''
        Permite generar una busqueda compleja de multiples valores en multiples campos
        por medio de sentencias OR
        campos: campo y tipo de consulta (ej. [nombre__icontains, apellido__icontains])
        valores: valores a buscar en la consulta OR separado por espacio (ej. pablo godoy) 
    '''
    q = Q()
    for campo in campos:
        for valor in valores:
            q |= Q(**{f'{campo}' : valor})
    return q

def app_installed(apps):
    '''
        Valida si la aplicacion esta instalada
    '''
    installed = list()
    for app in apps:
        if app in settings.INSTALLED_APPS:
            installed.append(apps[app])
    return installed

#
# FUNCIONES Y VISTAS GENERICAS
#    

def home(request):
    info = {
        'general': {
            'nombre_sitio': gConfiguracion.get_value('sitio', 'nombre')
        },
        'contenido': {
            'title': _('ERPv2'),
            'h1': _('Aplicaciones instaladas'),
        },
        'apps': app_installed(settings.INFORMACION_APLICACIONES),
        'opciones': {
            'ir': _('Ir'),
        },
    }
    return render(request, 'home.html', info)



class UsuarioLoginView(LoginView):
    template_name = "usuarios/login.html"
    extra_context = {
        'title': _('Página de ingreso'),
        'opciones': {
            'submit': _('Ingresar'),
            'reset': _('He olvidado la contraseña'),
        },
    }

    def get_context_data(self):
        context = super(UsuarioLoginView, self).get_context_data()
        context['general'] = {'nombre_sitio': gConfiguracion.get_value('sitio', 'nombre')}
        return context


class UsuarioLogoutView(LogoutView):
    pass


class UsuarioPasswordResetView(PasswordResetView):
    template_name = "usuarios/password_reset_form.html"
    subject_template_name = "usuarios/password_reset_subject.txt"
    email_template_name = "usuarios/password_reset_email.html"
    success_url = reverse_lazy("usuarios:password_reset_done")
    extra_context = {
        'opciones': {
            'submit': _('Soliciar'),
        },
    }

    def get_context_data(self):
        context = super(UsuarioPasswordResetView, self).get_context_data()
        context['general'] = {'nombre_sitio': gConfiguracion.get_value('sitio', 'nombre')}
        return context


class UsuarioPasswordResetDoneView(PasswordResetDoneView):
    template_name = "usuarios/password_reset_done.html"


class UsuarioPasswordResetConfirmView(PasswordResetConfirmView):
    template_name = "usuarios/password_reset_confirm.html"
    success_url = reverse_lazy("usuarios:password_reset_complete")
    extra_context ={
        'opciones': {
            'submit': _('Cambiar')
        },
    }

    def get_context_data(self):
        context = super(UsuarioPasswordResetConfirmView, self).get_context_data()
        context['general'] = {'nombre_sitio': gConfiguracion.get_value('sitio', 'nombre')}
        return context


class UsuarioPasswordResetCompleteView(PasswordResetCompleteView):
    template_name = "usuarios/password_reset_complete.html"
    extra_context ={
        'opciones': {
            'ingresar': _('Ingresar')
        },
    }

    def get_context_data(self):
        context = super(UsuarioPasswordResetCompleteView, self).get_context_data()
        context['general'] = {'nombre_sitio': gConfiguracion.get_value('sitio', 'nombre')}
        return context


class UsuarioPasswordChangeView(LoginRequiredMixin, SuccessMessageMixin, PasswordChangeView):
    template_name = 'usuarios/password_change_form.html'
    success_message = 'Contraseña cambiada correctamente'
    success_url = reverse_lazy('usuarios:perfil')
    extra_context = {
        'opciones': {
            'submit': _('Cambiar'),
        },
    }

    def get_context_data(self):
        context = super(UsuarioPasswordChangeView, self).get_context_data()
        context['general'] = {'nombre_sitio': gConfiguracion.get_value('sitio', 'nombre')}
        return context


class UsuarioPerfil(PersonalTemplateView):
    '''
        Muestra la información del perfil de usuario
    '''
    template_name = 'usuarios/perfil.html'
    permission_required = 'usuarios.view_perfil'
    extra_context ={
        'title': _('Perfil'),
    }

    def get_context_data(self):
        context = super(UsuarioPerfil, self).get_context_data()
        context['object'] = Usuario.objects.get(pk=self.request.user.id)
        return context


class UsuarioActualizar(PersonalUpdateView):
    '''
        Actualización del usuario que está logueado actualmente
        (Unicamente actualiza el propio perfil)
    '''
    permission_required = 'usuarios.change_perfil'
    model = Usuario
    fields = ['first_name', 'last_name', 'email']
    success_message = 'Usuario actualizado correctamente'
    success_url = reverse_lazy('usuarios:perfil')
    extra_context ={
        'title': _('Actualización de datos'),
        'opciones': {
            'submit': _('Modificar'),
        },
    }

    def get_object(self):
        return Usuario.objects.get(pk=self.request.user.id)

    def get_context_data(self, *args, **kwargs):
        context = super(UsuarioActualizar, self).get_context_data(*args, **kwargs)
        context['aditional_form'] = PerfilForm(instance=self.request.user.perfil)
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()
        aditional_form = PerfilForm(request.POST or None, instance=self.object.perfil)

        if form.is_valid() and aditional_form.is_valid():
            form.save()
            aditional_form.save()
            return HttpResponseRedirect(self.get_success_url())
        else:
            return super(UsuarioActualizar, self).form_invalid(form)


class UsuarioNuevoFormView(PersonalFormView):
    template_name = 'usuarios/usuario_form.html'
    permission_required = 'usuarios.add_usuario'
    form_class = CustomUserCreationForm
    success_message = _('Usuario creado correctamente')
    success_url = reverse_lazy('usuarios:home')
    extra_context ={
        'title': _('Creación de usuario'),
        'opciones': {
            'submit': _('Crear'),
        },
    }

    def form_valid(self, form):
        if form.is_valid():
            form.save()
        return super(UsuarioNuevoFormView, self).form_valid(form)


class UsuarioListView(PersonalListView):
    '''
        Lista de usuarios registrados en la aplicacion
    '''
    permission_required = 'usuarios.view_usuario'
    model = Usuario
    ordering = ('username')
    paginate_by = 12
    extra_context ={
        'title': _('Lista de usuarios'),
        'opciones': {
            'etiqueta': _('Opciones'),
            'editar': _('Editar'),
        },
    }

    def get_queryset(self):
        valor_busqueda = self.request.GET.get('valor')
        if valor_busqueda:
            if 'mail:' in valor_busqueda:
                return Usuario.objects.filter(email__icontains=valor_busqueda[5:]).order_by('email')
            elif 'name:' in valor_busqueda:
                return Usuario.objects.filter(
                    BusquedaNombres(
                        campos=['first_name__icontains', 'last_name__icontains'], 
                        valores=valor_busqueda[5:].split(' ')),
                    ).order_by('first_name', 'last_name')
            else:
                return Usuario.objects.filter(username__icontains=valor_busqueda).order_by('username')
        else:
            return super(UsuarioListView, self).get_queryset()


class UsuarioUpdateView(PersonalUpdateView):
    '''
        Actualización de usuarios terceros
        Como una funcion administrativa
    '''
    template_name = 'usuarios/usuario_form.html'
    permission_required = 'usuarios.change_usuario'
    model = Usuario
    form_class = CustomUserUpdateForm
    success_message = _('Usuario actualizado correctamente')
    success_url = reverse_lazy('usuarios:listar')
    extra_context ={
        'title': _('Actualizar Usuarios'),
        'opciones': {
            'submit': _('Modificar'),
        },
    }

    def get_context_data(self, *args, **kwargs):
        context = super(UsuarioUpdateView, self).get_context_data(*args, **kwargs)
        context['aditional_form'] = PerfilForm(instance=self.object.perfil)
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, instance=self.object)
        aditional_form = PerfilForm(request.POST, instance=self.object.perfil)

        if form.is_valid() and aditional_form.is_valid():
            form.save()
            aditional_form.save()
            return HttpResponseRedirect(self.get_success_url())
        else:
            return super(UsuarioUpdateView, self).form_invalid(form)


class RegionalizacionCreateView(PersonalFormView):
    '''
        Carga la informacion de departamentos/municipios para un país
    '''
    template_name = 'usuarios/regionalizacion_uploadform.html'
    permission_required = 'usuarios.add_regionalizacion'
    form_class = RegionalizacionUploadForm
    success_url = reverse_lazy('usuarios:listar_regionalizacion')
    success_message = _('Regionalización cargada correctamente')
    extra_context ={
        'title': _('Cargar Regionalización'),
        'opciones': {
            'submit': _('Guardar'),
        },
    }

    def form_valid(self, *args, **kwargs):
        nombre_pais = self.request.POST["pais"].upper()
        pais = Regionalizacion.objects.filter(nombre=nombre_pais, padre__isnull=True)
        
        if not pais:
            Regionalizacion.objects.create(nombre=nombre_pais, usuario=self.request.user)
            pais = Regionalizacion.objects.latest('id')
        else:
            pais = pais.latest('id')

        departamentos_list = list()
        municipios_list = list()
        
        departamentos, municipios = self._cargar_informacion()
        
        qryset_deptos = Regionalizacion.objects.filter(padre=pais)
        
        for departamento in departamentos:
            nombre = departamento[:60].upper()
            if not qryset_deptos.filter(nombre=nombre).exists():
                # si el departamento no ha sido ingresado se ingresa en la lista para crearlo
                departamentos_list.append(Regionalizacion(nombre=nombre, usuario=self.request.user, padre=pais))

        Regionalizacion.objects.bulk_create(departamentos_list)   
        qryset_municipios = Regionalizacion.objects.prefetch_related('self').filter(padre__padre=pais)
        
        for info_municipio in municipios:
            valor = info_municipio.upper().split(',')

            if not qryset_municipios.filter(nombre=valor[1][:60], padre__nombre=valor[0][:60]).exists():
                # si el departamento no ha sido ingresado se ingresa en la lista para crearlo
                municipios_list.append(Regionalizacion(nombre=valor[1][:60], usuario=self.request.user, 
                    padre=qryset_deptos.filter(nombre=valor[0][:60]).latest('id')))

        Regionalizacion.objects.bulk_create(municipios_list)

        return super(RegionalizacionCreateView, self).form_valid(*args, **kwargs)

    def _cargar_informacion(self):
        '''
            Carga la información sin repetidos
        '''
        departamentos = list()
        municipios = list()

        sheet = load_workbook(self.request.FILES['archivo']).active
        for linea in sheet.iter_rows(min_row=2):
            departamento_nombre = linea[0].value
            municipio_nombre = linea[1].value

            if not departamento_nombre in departamentos:
                departamentos.append(f'{departamento_nombre}')
            if not f'{departamento_nombre},{municipio_nombre}' in municipios:
                municipios.append(f'{departamento_nombre},{municipio_nombre}')

        return departamentos, municipios


class RegionalizacionListView(PersonalListView):
    template_name = 'usuarios/regionalizacion_list.html'
    permission_required = 'usuarios.view_regionalizacion'
    model = Regionalizacion
    extra_context ={
        'title': _('Listado de Regionalización'),
        'opciones': {
            'editar': _('Editar')
        },
    }

    def get_queryset(self):
        return None

    def get_context_data(self, *args, **kwargs):
        context = super(RegionalizacionListView, self).get_context_data(*args, **kwargs)
        context['object_list'] = Regionalizacion.objects.filter(padre__isnull=True).order_by('nombre')
        return context


class RegionalizacionUpdateView(PersonalUpdateView):
    permission_required = 'usuarios.change_regionalizacion'
    model = Regionalizacion
    fields = ('nombre', 'vigente')
    success_message = _('Actualización exitosa.')
    success_url = reverse_lazy('usuarios:listar_regionalizacion')
    extra_context ={
        'title': _('Actualización'),
        'opciones': {
            'submit': _('Actualizar')
        },
    }
